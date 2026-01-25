import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, BarChart, Reference

# Store Excel path in one place; file is created in C:\PythonFiles if missing
FINAL_XLSX_PATH = r"C:\\PythonFiles\\final.xlsx"


def ensure_workbook(path=FINAL_XLSX_PATH):
    """Load existing workbook or create a new one with headers."""
    os.makedirs(os.path.dirname(path), exist_ok=True)
    try:
        wb = load_workbook(path)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Temperature Data"
        ws.append(["Date", "Fahrenheit", "Celsius"])
        wb.save(path)
    return wb

# Function to convert temperature from Fahrenheit to Celsius
def convertData(temp_input):
    converted_value = (temp_input - 32) * 5.0/9.0
    return converted_value

# Function to get user input for temperature data and save to Excel file
def getInput():
    num_entries = int(input("How many entries are you entering? "))
    wb = ensure_workbook()
    ws = wb.active
    
    for i in range(num_entries):
        date_input = input("Enter the date: ")
        print()
        temp_input = float(input("Enter the temperature in fahrenheit: "))
        
        converted_value = convertData(temp_input)
        # Append row to Excel
        try:
            ws.append([date_input, temp_input, converted_value])
            wb.save(FINAL_XLSX_PATH)
            current_time = datetime.now()
            print(f"The following data was saved at {current_time}: {date_input}, {temp_input}, {converted_value}")
            print()
        except Exception as e:
            print(f"Failed to save data: {e}")
            print()


# Function to view data from Excel file
def viewData(file_path=FINAL_XLSX_PATH):
    try:
        print(f"Reading data from: {file_path}")
        print()
        wb = ensure_workbook(file_path)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            print(', '.join(str(cell) for cell in row))
    except FileNotFoundError:
        print(f"Error: The file '{file_path}' does not exist.")
    except Exception as e:
        print(f"Error reading file: {e}")


# Function to create chart from Excel data
def createChart(chart_type):
    try:
        # Ask user to choose data source
        print("Which data source would you like to use for the chart?")
        print("1. Fahrenheit (initial data)")
        print("2. Celsius (converted data)")
        data_choice = input("Enter your choice (1 or 2): ")
        
        # Determine which source column to use in the stored sheet
        if data_choice == '1':
            source_column = 2  # Fahrenheit column in stored data
            data_label = "Fahrenheit"
        elif data_choice == '2':
            source_column = 3  # Celsius column in stored data
            data_label = "Celsius"
        else:
            print("Invalid choice. Using Fahrenheit as default.")
            source_column = 2
            data_label = "Fahrenheit"
        
        wb = ensure_workbook()
        ws = wb.active
        dates = []
        temps = []
        # Skip header row (row 1)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None and row[source_column - 1] is not None:
                dates.append(row[0])
                temps.append(float(row[source_column - 1]))

        if not dates or not temps:
            print("No data available to chart. Please input data first.")
            return
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Temperature Data"
        
        # Add headers
        ws['A1'] = 'Date'
        ws['B1'] = data_label
        
        # Add data to worksheet
        for i, (date, temp) in enumerate(zip(dates, temps), start=2):
            ws[f'A{i}'] = date
            ws[f'B{i}'] = temp
        
        # Create chart based on chart_type
        if chart_type.lower() == 'line':
            chart = LineChart()
        elif chart_type.lower() == 'bar':
            chart = BarChart()
        else:
            print("Invalid chart type. Using bar chart as default.")
            chart = BarChart()
        
        # Set chart data and categories
        data = Reference(ws, min_col=2, min_row=1, max_row=len(temps) + 1)
        categories = Reference(ws, min_col=1, min_row=2, max_row=len(dates) + 1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        
        # Label axes
        chart.x_axis.title = "Date"
        chart.y_axis.title = f"Temperature ({data_label})"
        
        # Create title with student ID and current date
        current_date = datetime.now().strftime("%m/%d/%Y")
        chart.title = f"Jarand5122 {current_date}"
        
        # Add chart to worksheet
        ws.add_chart(chart, "D2")
        
        # Save workbook
        wb.save(FINAL_XLSX_PATH)
        print(f"Chart successfully created and saved to {FINAL_XLSX_PATH}")
        
    except FileNotFoundError:
        print(f"Error: The file '{FINAL_XLSX_PATH}' does not exist.")
    except Exception as e:
        print(f"Error creating chart: {e}")


# Function to generate report by asking user for chart type
def generateReport():
    try:
        print("Which type of chart would you like to create?")
        print("1. Line Chart")
        print("2. Bar Chart")
        chart_choice = input("Enter your choice (1 or 2): ")
        
        if chart_choice == '1':
            chart_type = 'line'
        elif chart_choice == '2':
            chart_type = 'bar'
        else:
            print("Invalid choice. Using bar chart as default.")
            chart_type = 'bar'
        
        # Call createChart function
        createChart(chart_type)
        
    except Exception as e:
        print(f"Error generating report: {e}")


print ("Jarand5122's Excel Spreadsheet Automation Menu")
print ("Choose a number from the following options:")
print ("1. Input Data")
print ("2. View Current Data")
print ("3. Generate Report")

user_choice = input()
if user_choice not in ['1', '2', '3']:
    print ("Invalid option selected. Please choose 1, 2, or 3.")
else:
    print (f"You selected option {user_choice} at {datetime.now()}")
    
if user_choice == '1':
    getInput()
elif user_choice == '2':
    viewData()
elif user_choice == '3':
    generateReport()
